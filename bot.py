import pandas as pd
from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes
import logging
import os
from datetime import datetime, timedelta
import pytz
from typing import Dict, Set
import sqlite3
import json
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
import io
import base64
import openpyxl
from openpyxl.utils import get_column_letter

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('/tmp/schedule_bot.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Конфигурация
TOKEN = os.getenv("BOT_TOKEN")
if not TOKEN:
    raise ValueError("BOT_TOKEN не задан")
ALLOWED_USERS = [int(id) for id in os.getenv("ALLOWED_USERS", "").split(",") if id]
DRIVE_CREDENTIALS = os.getenv("DRIVE_CREDENTIALS")
DRIVE_FOLDER_ID = os.getenv("DRIVE_FOLDER_ID")
DATA_DIR = "/tmp/data"
DB_FILE = "/tmp/schedule.db"
os.makedirs(DATA_DIR, exist_ok=True)

# Список преподавателей
TEACHER_NAMES = [
    "Tajieva A", "Mamirbaeva D", "Koyshekenova T", "Arzieva B", "Dauletmuratova X",
    "Jalgasov N", "Xudaybergenov A", "Allanazarova F", "Saparov S", "Balkibaeva V",
    "Jalgasbaeva G", "Bisenova A", "Kurbanbaeva D", "Tursunbaev B", "Yakupova K",
    "Xalmuratov I", "Tolepbergenov T", "Aymanova Sh", "Bayimbetova M", "Oringalieva D",
    "Qaypova B", "Dauletiyarova N", "Utebaeva A", "Bekbergenova G", "Kurbaniyazova S",
    "Xabibnazarova S", "Utepbergenova D", "Kanlibaeva E", "Kalimbetova K", "Elmuratova Z",
    "Esbergenova G", "Jiemuratova G", "Dauletbaeva N", "Joldasbaev O", "Narshabaeva A",
    "Utemisov A", "Bayniyazov A", "Abatov A", "Tleumuratova Z", "Naubetullaeva E", "Abdiev B",
    "Seitova Z", "Kurbanbaeva U", "Pazilov A", "Seytmuratov K", "Seytjanova U",
    "Kurbaniyazov M", "Madaminova N", "Esemuratova T", "Matmuratova G", "Qurbanbaeva D",
    "Atamuratova M"
]

# Столбцы групп для 1-го и 2-го курсов
GROUP_COLUMNS_FIRST_COURSE = [4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28]
GROUP_COLUMNS_SECOND_COURSE = [33, 35, 37, 39, 41, 43, 45, 47, 49, 51, 56, 58, 60, 62, 67, 69, 71, 73, 75, 77, 79, 81, 83, 85, 87]

# Диапазоны строк для дней недели
DAY_RANGES = ["5-16", "18-29", "31-42", "44-55", "57-68", "70-81"]
DAY_NAMES = ["DUYSEMBI", "SIYSHEMBI", "SARSHEMBI", "PIYSHEMBI", "JUMA", "SHEMBI"]

# Google Drive API
def init_drive():
    try:
        creds_json = base64.b64decode(DRIVE_CREDENTIALS).decode('utf-8')
        creds_dict = json.loads(creds_json)
        credentials = Credentials.from_service_account_info(creds_dict, scopes=['https://www.googleapis.com/auth/drive'])
        drive_service = build('drive', 'v3', credentials=credentials)
        logger.info("Google Drive initialized")
        return drive_service
    except Exception as e:
        logger.error(f"Google Drive init error: {str(e)}")
        raise

# Загрузка файла в Google Drive
def upload_to_drive(drive_service, file_path, folder_id):
    try:
        file_name = os.path.basename(file_path)
        file_metadata = {
            'name': file_name,
            'parents': [folder_id]
        }
        media = MediaFileUpload(file_path)
        file = drive_service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id'
        ).execute()
        logger.info(f"Uploaded {file_name} to Google Drive, ID: {file.get('id')}")
        return file.get('id')
    except Exception as e:
        logger.error(f"Upload to Drive error for {file_path}: {str(e)}")
        return None

# Скачивание последнего файла из Google Drive
def download_latest_from_drive(drive_service, folder_id, prefix):
    try:
        results = drive_service.files().list(
            q=f"'{folder_id}' in parents and name contains '{prefix}'",
            orderBy='createdTime desc',
            pageSize=1,
            fields="files(id, name)"
        ).execute()
        files = results.get('files', [])
        if not files:
            logger.warning(f"No files found with prefix {prefix}")
            return None
        file_id = files[0]['id']
        file_name = files[0]['name']
        file_path = os.path.join(DATA_DIR, file_name)
        request = drive_service.files().get_media(fileId=file_id)
        fh = io.FileIO(file_path, 'wb')
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            status, done = downloader.next_chunk()
        logger.info(f"Downloaded {file_name} from Google Drive to {file_path}")
        return file_path
    except Exception as e:
        logger.error(f"Download from Drive error for prefix {prefix}: {str(e)}")
        return None

# FastAPI приложение
from fastapi import FastAPI, Request
from contextlib import asynccontextmanager

app = FastAPI()
bot_app = None
drive_service = None

@asynccontextmanager
async def lifespan(app: FastAPI):
    global drive_service
    drive_service = init_drive()
    await init_bot()
    yield
    await bot_app.shutdown()

app = FastAPI(lifespan=lifespan)

# Инициализация базы данных
def init_db():
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute('''CREATE TABLE IF NOT EXISTS original_schedule
                       (id INTEGER PRIMARY KEY AUTOINCREMENT,
                        Kun TEXT, Jupliq INTEGER, Topar TEXT,
                        Pan TEXT, Oqitiwshi TEXT, Kabinet TEXT)''')
        cursor.execute('''CREATE TABLE IF NOT EXISTS changes_schedule
                       (id INTEGER PRIMARY KEY AUTOINCREMENT,
                        Kun TEXT, Jupliq INTEGER, Topar TEXT,
                        Pan TEXT, Oqitiwshi TEXT, Kabinet TEXT)''')
        cursor.execute('''CREATE TABLE IF NOT EXISTS users
                       (user_id INTEGER PRIMARY KEY,
                        role TEXT, teacher_name TEXT, group_name TEXT,
                        notifications INTEGER DEFAULT 1)''')
        conn.commit()
        logger.info("Database initialized successfully")
    except Exception as e:
        logger.error(f"Database initialization error: {str(e)}")
        raise
    finally:
        cursor.close()
        conn.close()

# Проверка существования таблицы
def check_table_exists(table_name: str) -> bool:
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
        exists = bool(cursor.fetchone())
        return exists
    except Exception as e:
        logger.error(f"Table check error for {table_name}: {str(e)}")
        return False
    finally:
        cursor.close()
        conn.close()

# Сохранение пользователей в базу и Google Drive
def save_user_to_db(user_id: int, user_data: 'UserData'):
    if not check_table_exists("users"):
        logger.warning("Users table does not exist, reinitializing database")
        init_db()
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute('''INSERT OR REPLACE INTO users
                       (user_id, role, teacher_name, group_name, notifications)
                       VALUES (?, ?, ?, ?, ?)''',
                       (user_id, user_data.role, user_data.teacher_name,
                        user_data.group, int(user_data.notifications)))
        conn.commit()
        logger.debug(f"User {user_id} saved to database")
        cursor.execute('SELECT * FROM users')
        users = [
            {'user_id': row[0], 'role': row[1], 'teacher_name': row[2],
             'group_name': row[3], 'notifications': bool(row[4])}
            for row in cursor.fetchall()
        ]
        json_path = os.path.join(DATA_DIR, 'users.json')
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(users, f, ensure_ascii=False, indent=2)
        upload_to_drive(drive_service, json_path, DRIVE_FOLDER_ID)
    except Exception as e:
        logger.error(f"User save error for {user_id}: {str(e)}")
    finally:
        cursor.close()
        conn.close()

# Загрузка пользователей из базы или Google Drive
def load_users_from_db() -> Dict[int, 'UserData']:
    users = {}
    json_path = os.path.join(DATA_DIR, 'users.json')
    try:
        downloaded = download_latest_from_drive(drive_service, DRIVE_FOLDER_ID, 'users.json')
        if downloaded and os.path.exists(downloaded):
            with open(downloaded, 'r', encoding='utf-8') as f:
                users_data = json.load(f)
            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()
            cursor.execute('DELETE FROM users')
            for user in users_data:
                cursor.execute('''INSERT OR REPLACE INTO users
                               (user_id, role, teacher_name, group_name, notifications)
                               VALUES (?, ?, ?, ?, ?)''',
                               (user['user_id'], user['role'], user['teacher_name'],
                                user['group_name'], int(user['notifications'])))
            conn.commit()
            cursor.close()
            conn.close()
            logger.info(f"Restored {len(users_data)} users from Google Drive")
    except Exception as e:
        logger.error(f"Failed to restore users from Drive: {str(e)}")
    if not check_table_exists("users"):
        logger.warning("Users table does not exist, skipping load")
        return users
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute('SELECT user_id, role, teacher_name, group_name, notifications FROM users')
        for row in cursor.fetchall():
            user_id, role, teacher_name, group_name, notifications = row
            user_data = UserData()
            user_data.role = role
            user_data.teacher_name = teacher_name
            user_data.group = group_name
            user_data.notifications = bool(notifications)
            users[user_id] = user_data
        logger.info(f"Loaded {len(users)} users from database")
    except Exception as e:
        logger.error(f"User load error: {str(e)}")
    finally:
        cursor.close()
        conn.close()
    return users

# Словарь групп
GROUP_UNIONS = {
    "101": ["101"], "102": ["102"], "103": ["103"], "104": ["104"], "105": ["105"],
    "106": ["106"], "107": ["107"], "108": ["108"], "109": ["109"], "110": ["110"],
    "111": ["111"], "112": ["112"], "113": ["113"], "201": ["201"], "202": ["202"],
    "203": ["203"], "204": ["204"], "205": ["205"], "206": ["206"], "207": ["207"],
    "208": ["208"], "209": ["209"], "210": ["210"], "301": ["301"], "303": ["303"],
    "304": ["304"], "305": ["305"], "4G": ["4G"], "4D": ["4D"], "4E": ["4E"], "4J": ["4J"],
    "4Z": ["4Z"], "4I": ["4I"], "4K": ["4K"],
    "101-102": ["101", "102"], "103-104": ["103", "104"], "105-106": ["105", "106"],
    "107-108": ["107", "108"], "109-110": ["109", "110"], "201-202": ["201", "202"],
    "201-206": ["201", "202", "203", "204", "205", "206"],
    "201-202-203-204-205-206": ["201", "202", "203", "204", "205", "206"],
    "207-208": ["207", "208"], "207-208-209": ["207", "208", "209"], "303-304": ["303", "304"],
    "4G-4J": ["4G", "4J"], "4Z-4I": ["4Z", "4I"], "4G-4D-4E-4J": ["4G", "4D", "4E", "4J"],
    "101-102-103-104-112": ["101", "102", "103", "104", "112"],
    "105-106-107-108-109-110-113": ["105", "106", "107", "108", "109", "110", "113"]
}

# Время пар
PAIR_TIMES = {
    1: "8:30-9:50", 2: "10:00-11:20", 3: "11:30-12:50",
    4: "13:00-14:20", 5: "14:30-15:50", 6: "16:00-17:20"
}

# Дни недели
DAYS_OF_WEEK = ["DUYSEMBI", "SIYSHEMBI", "SARSHEMBI", "PIYSHEMBI", "JUMA", "SHEMBI"]

# Клавиатуры
ROLE_KEYBOARD = ReplyKeyboardMarkup(
    [["Oqıtıwshı", "Student"]],
    one_time_keyboard=True,
    resize_keyboard=True
)
TEACHER_KEYBOARD = ReplyKeyboardMarkup(
    [["Búgin", "Erteń"], ["Kúndi tańlaw", "Tolıq hápteni kóriw"]],
    resize_keyboard=True
)
TEACHER_KEYBOARD_ADMIN = ReplyKeyboardMarkup(
    [["Búgin", "Erteń"], ["Kúndi tańlaw", "Tolıq hápteni kóriw"], ["Kesteni óshiriw"]],
    resize_keyboard=True
)
STUDENT_KEYBOARD = ReplyKeyboardMarkup(
    [["Búgin", "Erteń"], ["Kúndi tańlaw", "Tolıq hápteni kóriw"]],
    resize_keyboard=True
)
DAY_KEYBOARD = ReplyKeyboardMarkup(
    [DAYS_OF_WEEK[i:i+2] for i in range(0, len(DAYS_OF_WEEK), 2)] + [["Artqa qaytıw"]],
    resize_keyboard=True
)

class UserData:
    def __init__(self):
        self.role: str = None
        self.teacher_name: str = None
        self.group: str = None
        self.notifications: bool = True

class BotData:
    def __init__(self):
        self.original_file: str = None
        self.last_file: str = None
        self.users: Dict[int, UserData] = {}
        self.subscribed_users: Set[int] = set()

def save_to_db(df, table_name):
    try:
        conn = sqlite3.connect(DB_FILE)
        df.to_sql(table_name, conn, if_exists='replace', index=False)
        logger.debug(f"Data saved to {table_name}")
        csv_path = os.path.join(DATA_DIR, f"{table_name}.csv")
        df.to_csv(csv_path, index=False)
        upload_to_drive(drive_service, csv_path, DRIVE_FOLDER_ID)
    except Exception as e:
        logger.error(f"Database save error for {table_name}: {str(e)}")
    finally:
        conn.close()

def get_from_db(table_name):
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
        if not cursor.fetchone():
            csv_path = download_latest_from_drive(drive_service, DRIVE_FOLDER_ID, f"{table_name}.csv")
            if csv_path and os.path.exists(csv_path):
                df = pd.read_csv(csv_path)
                df.to_sql(table_name, conn, if_exists='replace', index=False)
                logger.info(f"Restored {table_name} from Google Drive")
                return df
            return pd.DataFrame()
        df = pd.read_sql(f'SELECT * FROM {table_name}', conn)
        logger.debug(f"Data retrieved from {table_name}: {len(df)} rows")
        return df
    except Exception as e:
        logger.debug(f"Database read for {table_name} returned empty due to: {str(e)}")
        return pd.DataFrame()
    finally:
        cursor.close()
        conn.close()

def clear_db(drive_service, folder_id):
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute("DROP TABLE IF EXISTS original_schedule")
        cursor.execute("DROP TABLE IF EXISTS changes_schedule")
        cursor.execute('''CREATE TABLE original_schedule
                       (id INTEGER PRIMARY KEY AUTOINCREMENT,
                        Kun TEXT, Jupliq INTEGER, Topar TEXT,
                        Pan TEXT, Oqitiwshi TEXT, Kabinet TEXT)''')
        cursor.execute('''CREATE TABLE changes_schedule
                       (id INTEGER PRIMARY KEY AUTOINCREMENT,
                        Kun TEXT, Jupliq INTEGER, Topar TEXT,
                        Pan TEXT, Oqitiwshi TEXT, Kabinet TEXT)''')
        conn.commit()
        logger.info("Database schedules cleared")
        
        # Удаление локальных файлов
        for file in os.listdir(DATA_DIR):
            if file.startswith(("keste_bot_", "Sabaq_keste_DELL_")):
                os.remove(os.path.join(DATA_DIR, file))
        logger.info("Local schedule and working files cleared")
        
        # Удаление файлов расписания на Google Drive
        try:
            query = f"'{folder_id}' in parents"
            results = drive_service.files().list(
                q=query,
                fields="files(id, name)",
                pageSize=100
            ).execute()
            files = results.get('files', [])
            deleted_count = 0
            for file in files:
                if file['name'].startswith(('keste_bot_', 'Sabaq_keste_DELL_', 'original_schedule', 'changes_schedule')):
                    drive_service.files().delete(fileId=file['id']).execute()
                    logger.info(f"Deleted Drive file: {file['name']} (ID: {file['id']})")
                    deleted_count += 1
            logger.info(f"Deleted {deleted_count} schedule files from Google Drive")
        except Exception as e:
            logger.error(f"Failed to delete Drive files: {str(e)}")
            return False
            
        return True
    except Exception as e:
        logger.error(f"Database clear error: {str(e)}")
        return False
    finally:
        cursor.close()
        conn.close()

# Функции для обработки рабочей таблицы
def create_column_set(columns):
    """Создаёт словарь для быстрого поиска столбцов групп."""
    return {str(col): True for col in columns}

def contains_teacher_name(cell_value, teacher_name):
    """Проверяет, содержит ли ячейка имя преподавателя."""
    if not cell_value:
        return False
    clean_cell_value = ''.join(cell_value.strip().split()).replace('.', '')
    clean_teacher_name = ''.join(teacher_name.strip().split()).replace('.', '')
    return clean_teacher_name.lower() in clean_cell_value.lower()

def get_group_list(ws, merge_area, group_col_set):
    """Получает список групп из объединённой ячейки."""
    groups = []
    start_col = merge_area[0][0].column
    end_col = start_col + merge_area[0][0].merge_area.width - 1
    for col in range(start_col, end_col + 1):
        if str(col) in group_col_set:
            group_name = ws.cell(row=3, column=col).value
            if group_name:
                groups.append(str(group_name))
    return groups

def get_union_name(group_list, group_unions):
    """Находит объединённое имя группы из словаря GROUP_UNIONS."""
    if not group_list:
        return ""
    group_set = set(group_list)
    for union_name, groups in GROUP_UNIONS.items():
        if set(groups) == group_set:
            return union_name
    return "-".join(sorted(group_list))

def get_audience(ws, row, last_col, group_columns):
    """Извлекает номер аудитории."""
    audience = ""
    if last_col + 1 <= ws.max_column:
        audience = ws.cell(row=row, column=last_col + 1).value or ""
    if not audience:
        for col in group_columns:
            if col > last_col:
                audience = ws.cell(row=row, column=col + 1).value or ""
                if audience:
                    break
    return audience if audience else "JOQ"

def process_course(ws, teacher_name, group_columns, time_column, schedule_data):
    """Обрабатывает расписание для одного курса."""
    group_col_set = create_column_set(group_columns)
    
    for day_idx, day_range in enumerate(DAY_RANGES):
        day = DAY_NAMES[day_idx]
        start_row, end_row = map(int, day_range.split("-"))
        
        for row in range(start_row, end_row + 1):
            time = ws.cell(row=row, column=time_column).value or "JOQ"
            
            for col in group_columns:
                cell = ws.cell(row=row + 1, column=col)
                if cell.is_merged:
                    merge_area = cell.merge_area
                    if col == merge_area[0][0].column:
                        cell_value = cell.value
                        if cell_value and contains_teacher_name(cell_value, teacher_name):
                            group_list = get_group_list(ws, merge_area, group_col_set)
                            group_name = get_union_name(group_list, GROUP_UNIONS)
                            subject = ws.cell(row=row

, column=merge_area[0][0].column).value or "JOQ"
                            audience = get_audience(ws, row, merge_area[0][0].column + merge_area[0][0].merge_area.width - 1, group_columns)
                            schedule_data.append({
                                "Oqitiwshi": teacher_name,
                                "Kun": day,
                                "Jupliq": time,
                                "Topar": group_name,
                                "Pan": subject,
                                "Kabinet": audience
                            })

def create_temp_schedule_file(df, temp_file_path):
    """Создаёт временный файл Excel в формате бота."""
    try:
        with pd.ExcelWriter(temp_file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='keste', index=False)
        logger.info(f"Temporary schedule file created: {temp_file_path}")
        return True
    except Exception as e:
        logger.error(f"Error creating temporary file {temp_file_path}: {str(e)}")
        return False

def process_working_schedule(file_path, temp_file_path):
    """Обрабатывает рабочую таблицу и создаёт временный файл."""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        schedule_data = []
        
        for teacher in TEACHER_NAMES:
            process_course(ws, teacher, GROUP_COLUMNS_FIRST_COURSE, 3, schedule_data)
            process_course(ws, teacher, GROUP_COLUMNS_SECOND_COURSE, 32, schedule_data)
        
        df = pd.DataFrame(schedule_data)
        if df.empty:
            logger.warning("No schedule data extracted from working schedule")
            return False
        df = df[["Kun", "Jupliq", "Topar", "Pan", "Oqitiwshi", "Kabinet"]]
        
        # Создаём временный файл
        if not create_temp_schedule_file(df, temp_file_path):
            return False
        return True
    except Exception as e:
        logger.error(f"Error processing working schedule: {str(e)}")
        return False

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    context.bot_data.setdefault('bot_data', BotData())
    bot_data: BotData = context.bot_data['bot_data']
    if not bot_data.users:
        bot_data.users = load_users_from_db()
        bot_data.subscribed_users = {uid for uid, ud in bot_data.users.items() if ud.notifications}
    if user_id not in bot_data.users:
        bot_data.users[user_id] = UserData()
        save_user_to_db(user_id, bot_data.users[user_id])
    else:
        bot_data.subscribed_users.add(user_id)
    user_data = bot_data.users[user_id]
    if user_data.role:
        keyboard = (TEACHER_KEYBOARD_ADMIN if user_id in ALLOWED_USERS else TEACHER_KEYBOARD) if user_data.role == "Oqıtıwshı" else STUDENT_KEYBOARD
        await update.message.reply_text(
            f"Salem! Siz {user_data.role} sıpatında dizimnen óttińiz. Sabaq kesteńizdi kóriw ushın túymeni tańlań:",
            reply_markup=keyboard
        )
    else:
        await update.message.reply_text(
            "Salem! Botqa xosh keldińiz. Óz rolińizdi tańlań:",
            reply_markup=ROLE_KEYBOARD
        )
    logger.info(f"User {user_id} started the bot")

async def notify_on(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    bot_data: BotData = context.bot_data.get('bot_data')
    if not bot_data or user_id not in bot_data.users:
        await update.message.reply_text("Iltimas, /start buyrıǵın baslań!")
        return
    user_data = bot_data.users[user_id]
    user_data.notifications = True
    bot_data.subscribed_users.add(user_id)
    save_user_to_db(user_id, user_data)
    await update.message.reply_text("Xabarlandırıwlar qosıldı! Erteń sabaq bolsa, sizge xabar jiberiledi.")
    logger.info(f"User {user_id} enabled notifications")

async def notify_off(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    bot_data: BotData = context.bot_data.get('bot_data')
    if not bot_data or user_id not in bot_data.users:
        await update.message.reply_text("Iltimas, /start buyrıǵın baslań!")
        return
    user_data = bot_data.users[user_id]
    user_data.notifications = False
    bot_data.subscribed_users.discard(user_id)
    save_user_to_db(user_id, user_data)
    await update.message.reply_text("Xabarlandırıwlar óshirildi! Endi sizge eskertiw jiberilmeydi.")
    logger.info(f"User {user_id} disabled notifications")

async def handle_role(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    role = update.message.text.strip()
    bot_data: BotData = context.bot_data.get('bot_data')
    if not bot_data or user_id not in bot_data.users:
        await update.message.reply_text("Iltimas, /start buyrıǵın baslań!")
        return
    if role in ["Oqıtıwshı", "Student"]:
        bot_data.users[user_id].role = role
        save_user_to_db(user_id, bot_data.users[user_id])
        reply_text = {
            "Oqıtıwshı": "Iltimas atı jónińizdi jazıń (máselen: Xalmuratov I):",
            "Student": "Toparıńızdı jazıń (máselen: 101,201,301, 4-kurslar: 4G):"
        }
        await update.message.reply_text(reply_text[role])
    else:
        await update.message.reply_text("Qátelik! Rol durıs kiritilmegen.")

async def handle_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    if user_id not in ALLOWED_USERS:
        await update.message.reply_text("Sizde fayldı júklew ruxsatı joq!")
        logger.warning(f"Unauthorized file upload attempt by {user_id}")
        return
    bot_data: BotData = context.bot_data.get('bot_data')
    if not bot_data:
        bot_data = BotData()
        bot_data.users = load_users_from_db()
        context.bot_data['bot_data'] = bot_data
    try:
        file_name = update.message.document.file_name
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if file_name == "keste_bot_orig.xlsx":
            file_path = os.path.join(DATA_DIR, f"keste_bot_orig_{timestamp}.xlsx")
            bot_data.original_file = file_path
            is_original = True
            table_name = "original_schedule"
            is_working_schedule = False
        elif file_name == "keste_bot_ozgeris.xlsx":
            file_path = os.path.join(DATA_DIR, f"keste_bot_ozgeris_{timestamp}.xlsx")
            is_original = False
            table_name = "changes_schedule"
            is_working_schedule = False
        elif file_name == "Sabaq keste_DELL.xlsx":
            file_path = os.path.join(DATA_DIR, f"Sabaq_keste_DELL_{timestamp}.xlsx")
            temp_file_path = os.path.join(DATA_DIR, f"keste_bot_orig_{timestamp}.xlsx")
            bot_data.original_file = temp_file_path
            is_original = True
            table_name = "original_schedule"
            is_working_schedule = True
        elif file_name == "Sabaq keste_DELL_ozgeris.xlsx":
            file_path = os.path.join(DATA_DIR, f"Sabaq_keste_DELL_ozgeris_{timestamp}.xlsx")
            temp_file_path = os.path.join(DATA_DIR, f"keste_bot_ozgeris_{timestamp}.xlsx")
            is_original = False
            table_name = "changes_schedule"
            is_working_schedule = True
        else:
            await update.message.reply_text("Qátelik! Fayl atı durıs emes. 'keste_bot_orig.xlsx', 'keste_bot_ozgeris.xlsx', 'Sabaq keste_DELL.xlsx' yamasa 'Sabaq keste_DELL_ozgeris.xlsx' bolıwı kerek.")
            return
        file = await update.message.document.get_file()
        await file.download_to_drive(file_path)
        bot_data.last_file = file_path if not is_working_schedule else temp_file_path
        logger.debug(f"File saved: {file_path}")
        
        if is_working_schedule:
            if not process_working_schedule(file_path, temp_file_path):
                await update.message.reply_text("Qátelik! Fayl óńdelmedi, mazmunı bo's yamasa qáte bar.")
                return
            file_path = temp_file_path  # Используем временный файл для дальнейшей обработки
        else:
            check_result = check_excel_file(file_path)
            if check_result != "OK":
                await update.message.reply_text(f"Fayl tekserildi, qátelik bar: {check_result}")
                return
        
        df = pd.read_excel(file_path, sheet_name="keste", engine="openpyxl")
        save_to_db(df, table_name)
        upload_to_drive(drive_service, file_path, DRIVE_FOLDER_ID)
        await update.message.reply_text("Fayl sátli júklendi hám tekserip shıǵıldı!")
        await notify_users(context, is_original, user_id)
    except Exception as e:
        await update.message.reply_text(f"Qátelik: {str(e)}")
        logger.error(f"File upload error: {str(e)}")

def check_excel_file(file_path: str) -> str:
    try:
        required_columns = ["Kun", "Jupliq", "Topar", "Pan", "Oqitiwshi", "Kabinet"]
        df = pd.read_excel(file_path, sheet_name="keste", engine="openpyxl")
        column_aliases = {
            "Jupliq": ["Jupliq", "Jupliq", "Пара", "Lesson"],
            "Kun": ["Kun", "Kún", "Day"],
            "Topar": ["Topar", "Group"],
            "Pan": ["Pan", "Pán", "Subject"],
            "Oqitiwshi": ["Oqitiwshi", "Oqıtıwshı", "Teacher"],
            "Kabinet": ["Kabinet", "Room"]
        }
        missing_cols = []
        rename_dict = {}
        for required_col in required_columns:
            found = False
            for alias in column_aliases[required_col]:
                if alias in df.columns:
                    rename_dict[alias] = required_col
                    found = True
                    break
            if not found:
                missing_cols.append(required_col)
        if missing_cols:
            return f"Baǵanalar kemis: {', '.join(missing_cols)}"
        df.rename(columns=rename_dict, inplace=True)
        df.to_excel(file_path, sheet_name="keste", index=False, engine="openpyxl")
        return "OK"
    except Exception as e:
        return f"Fayldı tekseriwde qátelik júz berdi: {str(e)}"

def parse_groups(group_str: str) -> list:
    if isinstance(group_str, str):
        return group_str.split('-')
    return [str(group_str)]

def get_current_day() -> str:
    uzb_tz = pytz.timezone('Asia/Tashkent')
    today = datetime.now(uzb_tz).weekday()
    return DAYS_OF_WEEK[today] if today < 6 else "SHEMBI"

def get_tomorrow_day() -> str:
    uzb_tz = pytz.timezone('Asia/Tashkent')
    today = datetime.now(uzb_tz)
    tomorrow = today + timedelta(days=1)
    return DAYS_OF_WEEK[tomorrow.weekday()] if tomorrow.weekday() < 6 else "SHEMBI"

async def delete_schedule(bot_data: BotData, drive_service, folder_id):
    if clear_db(drive_service, folder_id):
        bot_data.original_file = None
        bot_data.last_file = None
        return True
    return False

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    text = update.message.text.strip()
    bot_data: BotData = context.bot_data.get('bot_data')
    if not bot_data or user_id not in bot_data.users:
        await update.message.reply_text("Iltimas, /start buyrıǵın baslań!")
        return
    user_data = bot_data.users[user_id]
    try:
        required_columns = ["Kun", "Jupliq", "Topar", "Pan", "Oqitiwshi", "Kabinet"]
        df = get_from_db("changes_schedule")
        if df.empty:
            df = get_from_db("original_schedule")
        if df.empty:
            await update.message.reply_text("Sabaq kestesi ele júklenbegen! Iltimas, sabır etiń!")
            return
        missing_cols = [c for c in required_columns if c not in df.columns]
        if missing_cols:
            await update.message.reply_text(f"Qátelik! Kestede kerekli baǵanalar joq: {', '.join(missing_cols)}")
            logger.error(f"Missing columns in schedule: {missing_cols}")
            return
        df["Oqitiwshi"] = df["Oqitiwshi"].astype(str).str.strip()
        df["Topar"] = df["Topar"].astype(str).str.strip()
        if user_data.role == "Oqıtıwshı":
            if not user_data.teacher_name:
                user_data.teacher_name = text
                save_user_to_db(user_id, user_data)
                keyboard = TEACHER_KEYBOARD_ADMIN if user_id in ALLOWED_USERS else TEACHER_KEYBOARD
                await update.message.reply_text(
                    f"Siz {text} sıpatında dizimnen óttińiz. Sabaq kesteńizdi kóriw ushın, túymeni tańlań:",
                    reply_markup=keyboard
                )
                return
            schedule = df[df["Oqitiwshi"] == user_data.teacher_name]
            keyboard = TEACHER_KEYBOARD_ADMIN if user_id in ALLOWED_USERS else TEACHER_KEYBOARD
            if text == "Búgin":
                current_day = get_current_day()
                day_schedule = schedule[schedule["Kun"] == current_day]
                logger.info(f"Teacher {user_data.teacher_name} today schedule rows: {len(day_schedule)}")
                if day_schedule.empty:
                    await update.message.reply_text(f"Sizde búgin ({current_day}) sabaq joq.")
                else:
                    message = format_schedule(day_schedule, "Oqıtıwshı")
                    await update.message.reply_text(message, parse_mode="Markdown")
            elif text == "Erteń":
                tomorrow_day = get_tomorrow_day()
                day_schedule = schedule[schedule["Kun"] == tomorrow_day]
                logger.info(f"Teacher {user_data.teacher_name} tomorrow schedule rows: {len(day_schedule)}")
                if day_schedule.empty:
                    await update.message.reply_text(f"Sizde erteń ({tomorrow_day}) sabaq joq.")
                else:
                    message = format_schedule(day_schedule, "Oqıtıwshı")
                    await update.message.reply_text(message, parse_mode="Markdown")
            elif text == "Kúndi tańlaw":
                await update.message.reply_text("Kúndi tańlań:", reply_markup=DAY_KEYBOARD)
            elif text == "Tolıq hápteni kóriw":
                logger.info(f"Teacher {user_data.teacher_name} full week schedule rows: {len(schedule)}")
                if schedule.empty:
                    await update.message.reply_text("Sizde bul háptede sabaq joq.")
                else:
                    message = format_schedule(schedule, "Oqıtıwshı")
                    await update.message.reply_text(message, parse_mode="Markdown")
            elif text == "Kesteni óshiriw" and user_id in ALLOWED_USERS:
                if await delete_schedule(bot_data, drive_service, DRIVE_FOLDER_ID):
                    await update.message.reply_text("Keste sátli óshirildi, Google Drive hám tazılandı!", reply_markup=keyboard)
                else:
                    await update.message.reply_text("Qátelik: Keste óshirilmadi!")
            elif text in DAYS_OF_WEEK:
                day_schedule = schedule[schedule["Kun"] == text]
                logger.info(f"Teacher {user_data.teacher_name} day {text} schedule rows: {len(day_schedule)}")
                if day_schedule.empty:
                    await update.message.reply_text(f"Sizde {text} kúni sabaq joq.")
                else:
                    message = format_schedule(day_schedule, "Oqıtıwshı")
                    await update.message.reply_text(message, parse_mode="Markdown")
            elif text == "Artqa qaytıw":
                await update.message.reply_text("Bas betke qaytıldı:", reply_markup=keyboard)
            else:
                await update.message.reply_text("Keshirersiz! Siz basqa oqıtıwshınıń kestesin kóre almaysız!")
        elif user_data.role == "Student":
            if not user_data.group:
                if text not in GROUP_UNIONS:
                    await update.message.reply_text(f"Topar '{text}' tabılmadı. Durıs topar atın jazıń!")
                    return
                user_data.group = text
                save_user_to_db(user_id, user_data)
                await update.message.reply_text(
                    f"Siz {text} toparı sıpatında dizimnen óttińiz. Sabaq kesteńizdi kóriw ushın túymeni basıń:",
                    reply_markup=STUDENT_KEYBOARD
                )
                return
            if text in GROUP_UNIONS:
                relevant_groups = GROUP_UNIONS[text]
                schedule = df[df["Topar"].apply(lambda x: any(g in parse_groups(x) for g in relevant_groups))]
            else:
                relevant_groups = GROUP_UNIONS[user_data.group]
                schedule = df[df["Topar"].apply(lambda x: any(g in parse_groups(x) for g in relevant_groups))]
            if text == "Búgin":
                current_day = get_current_day()
                day_schedule = schedule[schedule["Kun"] == current_day]
                logger.info(f"Student group {user_data.group} today schedule rows: {len(day_schedule)}")
                if day_schedule.empty:
                    await update.message.reply_text(f"Sizde búgin ({current_day}) sabaq joq.")
                else:
                    message = format_schedule(day_schedule, "Student")
                    await update.message.reply_text(message, parse_mode="Markdown")
            elif text == "Erteń":
                tomorrow_day = get_tomorrow_day()
                day_schedule = schedule[schedule["Kun"] == tomorrow_day]
                logger.info(f"Student group {user_data.group} tomorrow schedule rows: {len(day_schedule)}")
                if day_schedule.empty:
                    await update.message.reply_text(f"Sizde erteń ({tomorrow_day}) sabaq joq.")
                else:
                    message = format_schedule(day_schedule, "Student")
                    await update.message.reply_text(message, parse_mode="Markdown")
            elif text == "Kúndi tańlaw":
                await update.message.reply_text("Kúndi tańlań:", reply_markup=DAY_KEYBOARD)
            elif text == "Tolıq hápteni kóriw":
                logger.info(f"Student group {user_data.group} full week schedule rows: {len(schedule)}")
                if schedule.empty:
                    await update.message.reply_text("Sizde bul háptede sabaq joq.")
                else:
                    message = format_schedule(schedule, "Student")
                    await update.message.reply_text(message, parse_mode="Markdown")
            elif text in DAYS_OF_WEEK:
                day_schedule = schedule[schedule["Kun"] == text]
                logger.info(f"Student group {user_data.group} day {text} schedule rows: {len(day_schedule)}")
                if day_schedule.empty:
                    await update.message.reply_text(f"Sizde {text} kúni sabaq joq.")
                else:
                    message = format_schedule(day_schedule, "Student")
                    await update.message.reply_text(message, parse_mode="Markdown")
            elif text == "Artqa qaytıw":
                await update.message.reply_text("Bas betke qaytıldı:", reply_markup=STUDENT_KEYBOARD)
            elif text in GROUP_UNIONS:
                logger.info(f"Student viewing group {text} schedule rows: {len(schedule)}")
                if schedule.empty:
                    await update.message.reply_text(f"{text} toparında sabaq joq.")
                else:
                    message = format_schedule(schedule, "Student")
                    await update.message.reply_text(message, parse_mode="Markdown")
    except Exception as e:
        await update.message.reply_text(f"Qátelik: {str(e)}")
        logger.error(f"Message handling error: {str(e)}")

def format_schedule(df: pd.DataFrame, role: str) -> str:
    days_order = DAYS_OF_WEEK
    df = df.copy()
    df["Kun"] = pd.Categorical(df["Kun"], categories=days_order, ordered=True)
    df = df.sort_values(["Kun", "Jupliq"])
    message = ""
    current_day = None
    for _, row in df.iterrows():
        if row["Kun"] != current_day:
            if current_day:
                message += "\n============\n\n"
            message += f"📅 **{row['Kun']}**\n\n"
            current_day = row["Kun"]
        pair_time = PAIR_TIMES.get(row["Jupliq"], "Waqıt kórsetilmegen")
        if role == "Oqıtıwshı":
            message += (
                f"🕒 {row['Jupliq']}-Jupliq, {pair_time}\n"
                f"👤 Topar: {row['Topar']}\n"
                f"📚 Pán: {row['Pan']}\n"
                f"🚪 Kabinet: {row['Kabinet']}\n"
                "----------\n"
            )
        else:
            message += (
                f"🕒 {row['Jupliq']}-Jupliq, {pair_time}\n"
                f"📚 Pán: {row['Pan']}\n"
                f"👤 Oqıtıwshı: {row['Oqitiwshi']}\n"
                f"🚪 Kabinet: {row['Kabinet']}\n"
                "----------\n"
            )
    return message if message else "Sabaq joq"

async def notify_users(context: ContextTypes.DEFAULT_TYPE, is_original: bool, uploader_id: int):
    bot_data: BotData = context.bot_data.get('bot_data')
    if not bot_data:
        logger.error("Bot data not initialized")
        return
    try:
        logger.info(f"Starting notifications for is_original={is_original}, uploader_id={uploader_id}")
        logger.debug(f"Subscribed users: {bot_data.subscribed_users}")
        if is_original:
            message = "Xabarlandırıw! Jańa keste júklendi. Sabaq kesteńizdi kóriwińiz soraladı."
            if uploader_id in ALLOWED_USERS:
                await context.bot.send_message(chat_id=uploader_id, text=message)
                logger.info(f"Sent new schedule notification to admin {uploader_id}")
            for user_id in bot_data.subscribed_users:
                if user_id != uploader_id:
                    await context.bot.send_message(chat_id=user_id, text=message)
                    logger.debug(f"Sent new schedule notification to user {user_id}")
            logger.info(f"Sent new schedule notifications to {len(bot_data.subscribed_users)} users")
            return
        old_df = get_from_db("original_schedule")
        new_df = get_from_db("changes_schedule")
        if old_df.empty or new_df.empty:
            logger.warning("One of the schedules is empty, cannot compare")
            return
        old_df["Oqitiwshi"] = old_df["Oqitiwshi"].astype(str).str.strip()
        old_df["Topar"] = old_df["Topar"].astype(str).str.strip()
        new_df["Oqitiwshi"] = new_df["Oqitiwshi"].astype(str).str.strip()
        new_df["Topar"] = new_df["Topar"].astype(str).str.strip()
        changes = pd.concat([old_df, new_df]).drop_duplicates(keep=False)
        if changes.empty:
            logger.info("No changes detected in schedule")
            return
        affected_teachers = set(changes['Oqitiwshi'].dropna().unique())
        affected_groups = set()
        for group_str in changes['Topar'].dropna().unique():
            affected_groups.update(parse_groups(group_str))
        notified_users = set()
        for user_id, user_data in bot_data.users.items():
            if not user_data.notifications and user_id not in ALLOWED_USERS:
                continue
            if user_id in notified_users:
                continue
            if user_data.role == "Oqıtıwshı" and user_data.teacher_name in affected_teachers:
                await context.bot.send_message(
                    chat_id=user_id,
                    text="Xabarlandırıw! Kesteńizde ózgeris boldı. Iltimas tekserip kóriń."
                )
                notified_users.add(user_id)
                logger.debug(f"Sent change notification to teacher {user_id}")
            elif user_data.role == "Student" and user_data.group in GROUP_UNIONS:
                relevant_groups = GROUP_UNIONS[user_data.group]
                if any(group in affected_groups for group in relevant_groups):
                    await context.bot.send_message(
                        chat_id=user_id,
                        text="Xabarlandırıw! Kesteńizde ózgeris boldı. Iltimas tekserip kóriń."
                    )
                    notified_users.add(user_id)
                    logger.debug(f"Sent change notification to student {user_id}")
        if uploader_id in ALLOWED_USERS and uploader_id not in notified_users:
            await context.bot.send_message(
                chat_id=uploader_id,
                text="Xabarlandırıw! Kesteńizde ózgeris boldı. Iltimas tekserip kóriń."
            )
            logger.info(f"Sent change notification to admin {uploader_id}")
        logger.info(f"Sent change notifications to {len(notified_users)} users")
    except Exception as e:
        logger.error(f"Notification processing error: {str(e)}")

# Webhook endpoint
@app.post("/{token}")
async def webhook(token: str, request: Request):
    if token != TOKEN:
        logger.warning("Invalid webhook token received")
        return {"status": "error", "message": "Invalid token"}
    try:
        json_data = await request.json()
        update = Update.de_json(json_data, bot_app.bot)
        await bot_app.process_update(update)
        return {"status": "ok"}
    except Exception as e:
        logger.error(f"Webhook error: {str(e)}")
        return {"status": "error"}

# Root endpoint
@app.get("/")
async def root():
    return {"message": "Schedule Bot is running"}

# Инициализация бота
async def init_bot():
    global bot_app
    init_db()
    bot_app = Application.builder().token(TOKEN).build()
    bot_app.bot_data['bot_data'] = BotData()
    bot_data = bot_app.bot_data['bot_data']
    bot_data.users = load_users_from_db()
    bot_data.subscribed_users = {
        uid for uid, ud in bot_data.users.items() if ud.notifications
    }
    try:
        bot_data.original_file = download_latest_from_drive(drive_service, DRIVE_FOLDER_ID, "keste_bot_orig")
        bot_data.last_file = bot_data.original_file
        if bot_data.original_file:
            df = pd.read_excel(bot_data.original_file, sheet_name="keste", engine="openpyxl")
            save_to_db(df, "original_schedule")
        changes_file = download_latest_from_drive(drive_service, DRIVE_FOLDER_ID, "keste_bot_ozgeris")
        if changes_file:
            df = pd.read_excel(changes_file, sheet_name="keste", engine="openpyxl")
            save_to_db(df, "changes_schedule")
            bot_data.last_file = changes_file
    except Exception as e:
        logger.error(f"Failed to restore schedules: {str(e)}")
    bot_app.add_handler(CommandHandler("start", start))
    bot_app.add_handler(CommandHandler("notify_on", notify_on))
    bot_app.add_handler(CommandHandler("notify_off", notify_off))
    bot_app.add_handler(MessageHandler(filters.Regex(r"^(Oqıtıwshı|Student)$"), handle_role))
    bot_app.add_handler(MessageHandler(filters.Document.ALL, handle_file))
    bot_app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    await bot_app.initialize()
    await bot_app.start()
    logger.info("Bot initialized")

if __name__ == "__main__":
    import uvicorn
    init_db()
    uvicorn.run(app, host="0.0.0.0", port=int(os.getenv("PORT", 10000)))
